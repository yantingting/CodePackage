# coding: utf-8
"""
Python 3.6.0
uku周报
"""
# import sys
# import os
#
# sys.path.append("C:\\Users\\Mint\\Documents\\repos\\genie")

import datetime
import os
import pandas as pd
import numpy as np
import sys

sys.path.append('C:/workspace/repos/genie')

from utils3.data_io_utils import *

"""放款量、合同金额---------------------------------------------------------------------"""
fangkuan_sql="""
select
	case
		when effective_date <= current_date-8 then 'last_week'
		else 'this_week' end as week,
		return_flag,
		count(distinct id) as num,
		sum(approved_principal)/ 2000::numeric as prin
	from
		public.dw_gocash_go_cash_loan_gocash_core_loan
	where
		effective_date >= current_date-14
        and effective_date <= current_date-1
		and loan_status not in ('DENIED',
		'RESCIND',
		'APPROVING',
		'CREATED')
	group by
		1,
		2
"""
fangkuan_data=get_df_from_pg(fangkuan_sql)

import numpy as np
np.set_printoptions(suppress=False)

fangkuan_sum=fangkuan_data.groupby('week').sum()
fangkuan_sum.sort_values(by='week',inplace=True,ascending=False)
fangkuan_sum_T=fangkuan_sum.T

##新客户
fangkuan_new=fangkuan_data[fangkuan_data['return_flag']=='false']
fangkuan_new_sum=fangkuan_new.groupby('week').sum()
fangkuan_new_sum.sort_values(by='week',inplace=True,ascending=False)
fangkuan_new_sum_T=fangkuan_new_sum.T
#老客户
fangkuan_old=fangkuan_data[fangkuan_data['return_flag']=='true']
fangkuan_old_sum=fangkuan_old.groupby('week').sum()
fangkuan_old_sum.sort_values(by='week',inplace=True,ascending=False)
fangkuan_old_sum_T=fangkuan_old_sum.T

"""机审、人审通过率--------------------------------------------------------------------------"""

check_sql="""
select
    case when a.apply_time::date>=current_date-14 and a.apply_time::date<=current_date-8 then 'last_week' else 'this_week' end as week,
	a.return_flag,
	count(distinct case when a.apply_time ::date >= '2020-01-01' then a.id end) as application ,
	count(distinct case when a.apply_time ::date >= '2020-01-01' and risklevel is not null and risklevel <> '' then a.id end) as RC ,
	count(distinct case when a.apply_time ::date >= '2020-01-01' and risklevel in ('N', 'P') then a.id end) as RC_pass ,
	count(distinct case when a.apply_time ::date >= '2020-01-01' and bank_card_result in('BIND', 'BINDING') and photo_result = 'ENABLE' and risklevel in ('N', 'P') then a.id end) as bankandphoto_pass,
	count(distinct case when effective_date >= '2020-01-01' and loan_status not in ('DENIED', 'RESCIND', 'APPROVING', 'CREATED') then a.id else null end)as fangkuan
from
	(
	select
		*
	from
		public.dw_gocash_go_cash_loan_gocash_core_loan
	where
		apply_time ::date >= current_date-14
        and apply_time ::date <= current_date-1) a
left join (
	select
		*
	from
		rt_risk_mongo_gocash_installmentriskcontrolresult
	where
		(pipelinename like '%android_uku_new_v4_rctotal%' or pipelinename like '%android_uku_new_v5_rctotal%' or pipelinename like '%android_uku_old_rctotal%' or pipelinename like '%ios_new_uku_rctotal%' or pipelinename like '%ios_uku_old_rctotal%')
		and businessid = 'uku')b on
	a.id::varchar = b.loanid
group by
	1,
	2
"""
check_data=get_df_from_pg(check_sql)

check_data['rcpass_rate']=check_data['rc_pass']/check_data['rc']
check_data['pbpass_rate']=check_data['bankandphoto_pass']/check_data['rc_pass']
check_data['allpass_rate']=check_data['fangkuan']/check_data['application']
check_data.sort_values(by='week',inplace=True,ascending=False)
check_data_T=check_data.T


"""新用户区分降额通过率--------------------------------------------------------------------------"""
check_new_sql="""
select
    case when a.apply_time::date>=current_date-14 and a.apply_time::date<=current_date-8 then 'last_week' else 'this_week' end as week,
	count(distinct case when a.apply_time ::date >= '2020-01-01' then a.id end) as application ,
	count(distinct case when effective_date >= '2020-01-01' and loan_status not in ('DENIED', 'RESCIND', 'APPROVING', 'CREATED') then a.id else null end)as fangkuan,
	count(distinct case when effective_date >= '2020-01-01' and initial_principal= '500000' and loan_status not in ('DENIED', 'RESCIND', 'APPROVING', 'CREATED') then a.id else null end)as je_fangkuan,
    count(distinct case when effective_date >= '2020-01-01' and (initial_principal<> '500000' or initial_principal is null) and loan_status not in ('DENIED', 'RESCIND', 'APPROVING', 'CREATED') then a.id else null end)as std_fangkuan
from
	(
	select
		*
	from
		public.dw_gocash_go_cash_loan_gocash_core_loan
	where
		apply_time ::date >= current_date-14
        and apply_time ::date <= current_date-1
        and return_flag='false') a
left join (
	select
		loanid,
        initial_principal
    from
		rt_risk_mongo_gocash_installmentriskcontrolresult) b
    on
	a.id::varchar = b.loanid
group by
	1
"""
check_new_data=get_df_from_pg(check_new_sql)

check_new_data['je_pass_rate']=check_new_data['je_fangkuan']/check_new_data['application']
check_new_data['std_pass_rate']=check_new_data['std_fangkuan']/check_new_data['application']

check_new_data.sort_values(by='week',inplace=True,ascending=False)
check_new_data_T=check_new_data.T


"""首逾--------------------------------------------------------------------------"""
dpd1_sql="""
select
	case
		when effective_date <= current_date-23 then 'last_week'
		else 'this_week' end as week,
		return_flag,
		count(distinct case when late_date- effective_date = approved_period then loan_id else null end)/ count(distinct loan_id)::numeric dpd1
	from
		(
		select
			t1.id loan_id,
			effective_date,
			return_flag,
			loan_status,
			approved_period,
			late_date,
			case
				when t1.loan_status = 'COLLECTION' then current_date-t1.due_date + 1
				else 0 end as DPD,
				t1.approved_principal,
            initial_principal
			from
				dw_gocash_go_cash_loan_gocash_core_loan t1
			left join (
				select
					*
				from
					(
					select
						order_id,
						late_date,
						row_number() over(partition by order_id
					order by
						late_date asc) as num
					from
						dw_gocash_gocash_collection_col_case
					where
						order_status = 'COLLECTION_PAIDOFF'
						or order_status = 'COLLECTION')t
				where
					num = 1) t2 on
				t1.id = t2.order_id
            left join
            (select loanid,initial_principal from rt_risk_mongo_gocash_installmentriskcontrolresultgray)t3
            on t1.id::varchar=t3.loanid
            )t
	where
		effective_date is not null
		and effective_date >= current_date-29
		and effective_date <= current_date-16
		and loan_status not in ('DENIED',
		'RESCIND',
		'APPROVING',
		'CREATED')
        and (initial_principal<>'500000' or initial_principal is null)
	group by
		1,
		2
"""
dpd1_data=get_df_from_pg(dpd1_sql)

dpd1_data['dpd1']=dpd1_data['dpd1'].apply(lambda x: '%.2f%%' % (x*100))
dpd1_data_new_user=dpd1_data[dpd1_data['return_flag']=='false']
dpd1_data_old_user=dpd1_data[dpd1_data['return_flag']=='true']

dpd1_data_new_user.sort_values(by='week',inplace=True,ascending=False)
dpd1_data_new_user_T=dpd1_data_new_user.T
dpd1_data_old_user.sort_values(by='week',inplace=True,ascending=False)
dpd1_data_old_user_T=dpd1_data_old_user.T


"""DPD7--------------------------------------------------------------------------"""
dpd7_sql="""
select
	case
		when effective_date <= current_date-30 then 'last_week'
		else 'this_week' end as week,
		return_flag,
		count(distinct case when DPD >= 7 then loan_id else null end)/ count(distinct loan_id)::numeric dpd7
	from
		(
		select
		t1.id loan_id,
		effective_date,
		return_flag,
		loan_status,
		approved_period,
		late_date,
		approved_principal,
        initial_principal,
		case
			when loan_status = 'COLLECTION' then current_date-late_date
			else round(late_fee /(approved_principal*0.025))::int end as DPD
		from
			(
			select
				*
			from
				dw_gocash_go_cash_loan_gocash_core_loan
			where
				effective_date is not null
				and effective_date >= '2019-01-01'
				and loan_status not in ('DENIED',
				'RESCIND',
				'APPROVING',
				'CREATED')) t1
		left join (
			select
				*
			from
				(
				select
					order_id,
					late_date,
					row_number() over(partition by order_id
				order by
					late_date asc) as num
				from
					dw_gocash_gocash_collection_col_case
				where
					(order_status = 'COLLECTION_PAIDOFF'
					or order_status = 'COLLECTION')
					and app_id not in ('Credits','KASANDAAI','kasanda'))t
			where
				num = 1) t2 on
			t1.id = t2.order_id
            left join
            (select loanid,initial_principal from rt_risk_mongo_gocash_installmentriskcontrolresultgray)t3
            on t1.id::varchar=t3.loanid
            )t
	where
		effective_date is not null
		and effective_date >= current_date-36
		and effective_date <= current_date-23
		and loan_status not in ('DENIED',
		'RESCIND',
		'APPROVING',
		'CREATED')
        and (initial_principal<>'500000' or initial_principal is null)
	group by
		1,
		2
"""
dpd7_data=get_df_from_pg(dpd7_sql)

dpd7_data['dpd7']=dpd7_data['dpd7'].apply(lambda x: '%.2f%%' % (x*100))
dpd7_data_new_user=dpd7_data[dpd7_data['return_flag']=='false']
dpd7_data_old_user=dpd7_data[dpd7_data['return_flag']=='true']

dpd7_data_new_user.sort_values(by='week',inplace=True,ascending=False)
dpd7_data_new_user_T=dpd7_data_new_user.T
dpd7_data_old_user.sort_values(by='week',inplace=True,ascending=False)
dpd7_data_old_user_T=dpd7_data_old_user.T


"""5.出催情况--------------------------------------------------------------------------"""
cuihui_sql="""
select
	case
		when u.late_date <= current_date-11 then 'last_week'
		else 'this_week' end as week,
		u.return_flag,
		sum(late3_pi) / sum(case_num)::numeric as late3_pi_rate,
		sum(late3_i) / sum(case_num)::numeric as late3_i_rate
	from
		(
		select
			t1.late_date late_date,
			return_flag,
			count(distinct t1.order_id) as case_num,
			round(sum(approved_principal)/ 2100) as prin_sum
		from
			(
			select
				late_date,
				id,
				order_id,
				approved_principal,
				row_number() over(partition by order_id
			order by
				late_date desc) as num
			from
				dw_gocash_gocash_collection_col_case
			where
				order_status != 'ABNORMAL'
				and order_status != 'PAIDOFF'
				and app_id not in ('Credits','KASANDAAI','kasanda'))t1
		left join (
			select
				id,
				return_flag
			from
				dw_gocash_go_cash_loan_gocash_core_loan)t2 on
			t1.order_id = t2.id
        left join
            (select loanid,initial_principal from rt_risk_mongo_gocash_installmentriskcontrolresultgray)t3
            on t1.id::varchar=t3.loanid
		where
			t1.late_date >= current_date-17
			and t1.late_date <= current_date-4
			and num = 1
            and (initial_principal<>'500000' or initial_principal is null)
		group by
			late_date,
			return_flag
		order by
			late_date asc)u
	left join (
		select
			t1.late_date late_date,
			return_flag,
			count(case when t2.late_day<4 and refund_type = 'PRINCIPAL_INTEREST' then 1 else null end ) as late3_pi,
			count(case when t2.late_day<4 and refund_type = 'INTEREST' then 1 else null end ) as late3_i
		from
			(
			select
				late_date,
				id,
				order_id,
				approved_principal,
				row_number() over(partition by order_id
			order by
				late_date desc) as num
			from
				dw_gocash_gocash_collection_col_case
			where
				order_status != 'ABNORMAL'
				and order_status != 'PAIDOFF'
				and app_id not in ('Credits','KASANDAAI','kasanda'))t1
		inner join dw_gocash_gocash_collection_col_repayment t2 on
			t1.id = t2.case_id
		left join (
			select
				id,
				return_flag
			from
				dw_gocash_go_cash_loan_gocash_core_loan)t3 on
			t1.order_id = t3.id
        left join
            (select loanid,initial_principal from rt_risk_mongo_gocash_installmentriskcontrolresultgray)t4
            on t1.order_id::varchar=t4.loanid
		where
			t1.late_date >= current_date-17
			and t1.late_date <= current_date-4
			and num = 1
            and (initial_principal<>'500000' or initial_principal is null)
		group by
			t1.late_date,
			return_flag
		order by
			late_date asc)v on
		u.late_date = v.late_date
		and u.return_flag = v.return_flag
	group by
		1,
		2
"""
cuihui_data=get_df_from_pg(cuihui_sql)

cuihui_data=cuihui_data[cuihui_data['late3_pi_rate']>0]
cuihui_data['late3_pi_rate']=cuihui_data['late3_pi_rate'].apply(lambda x: '%.2f%%' % (x*100))
cuihui_data['late3_i_rate']=cuihui_data['late3_i_rate'].apply(lambda x: '%.2f%%' % (x*100))
cuihui_data_new_user=cuihui_data[cuihui_data['return_flag']=='false']
cuihui_data_old_user=cuihui_data[cuihui_data['return_flag']=='true']

cuihui_data_new_user.sort_values(by='week',inplace=True,ascending=False)
cuihui_data_new_user_T=cuihui_data_new_user.T
cuihui_data_old_user.sort_values(by='week',inplace=True,ascending=False)
cuihui_data_old_user_T=cuihui_data_old_user.T


"""复贷率--------------------------------------------------------------------------"""
fudai_sql="""
select
	case
		when day ::date >= current_date-37
		and day ::date <= current_date-31 then 'this_week'
		when day ::date >= current_date-44
		and day ::date <= current_date-38 then 'last_week'
		else 'recent_week' end as week,
	count(distinct customer_id) as new_all,
	count(distinct old_cus) fudai,
	count(distinct case when loan_status like '%PAIDOFF%' then customer_id end) as new_all_pay
from
	(
	select
		a.*,
		to_char(a.effective_date, 'YYYY-MM-DD') as day,
		b.effective_date as old,
		b.customer_id as old_cus
	from
		(
		select
			customer_id,
			effective_date,
			loan_status
		from
			dw_gocash_go_cash_loan_gocash_core_loan
		where
			effective_date >=current_date-44
			and return_flag = 'false')a
	left join (
		select
			*
		from
			dw_gocash_go_cash_loan_gocash_core_loan
		where
			effective_date >= current_date-44
			and return_flag = 'true')b on
		a.customer_id = b.customer_id)c
group by
	1
"""
fudai_data=get_df_from_pg(fudai_sql)

fudai_sql_clean=fudai_data[fudai_data['week']!='recent_week']
fudai_sql_clean['new_fudai']=fudai_sql_clean['fudai']/fudai_sql_clean['new_all']
fudai_sql_clean['new_pay_fudai']=fudai_sql_clean['fudai']/fudai_sql_clean['new_all_pay']
fudai_sql_clean['new_fudai']=fudai_sql_clean['new_fudai'].apply(lambda x: '%.2f%%' % (x*100))
fudai_sql_clean['new_pay_fudai']=fudai_sql_clean['new_pay_fudai'].apply(lambda x: '%.2f%%' % (x*100))
fudai_sql_clean.sort_values(by='week',inplace=True,ascending=False)
fudai_sql_clean_T=fudai_sql_clean.T

"""输出excel--------------------------------------------------------------------------"""
import pandas as pd
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'test'

ws.cell(1,2,'本周实际')
ws.cell(1,3,'上周表现')
ws.cell(2,1,'总指标')
ws.cell(3,1,'总合同金额')
ws.cell(3,2,fangkuan_sum_T.iloc[1,0])
ws.cell(3,3,fangkuan_sum_T.iloc[1,1])
ws.cell(4,1,'新客户指标(包括降额客户)')
ws.cell(5,1,'新客户申请笔数')
ws.cell(5,2,check_data_T.iloc[2,0])
ws.cell(5,3,check_data_T.iloc[2,2])
ws.cell(6,1,'机审通过率')
ws.cell(6,2,check_data_T.iloc[7,0])
ws.cell(6,3,check_data_T.iloc[7,2])
ws.cell(7,1,'人审通过率')
ws.cell(7,2,check_data_T.iloc[8,0])
ws.cell(7,3,check_data_T.iloc[8,2])
ws.cell(8,1,'新客户放款笔数')
ws.cell(8,2,fangkuan_new_sum_T.iloc[0,0])
ws.cell(8,3,fangkuan_new_sum_T.iloc[0,1])
ws.cell(9,1,'新客户合同金额')
ws.cell(9,2,fangkuan_new_sum_T.iloc[1,0].astype(int))
ws.cell(9,3,fangkuan_new_sum_T.iloc[1,1].astype(int))
from decimal import Decimal
ws.cell(10,1,'新客户总通过率')
ws.cell(10,2).value=ws.cell(8,2).value.astype(Decimal)/ws.cell(5,2).value
ws.cell(10,3).value=ws.cell(8,3).value.astype(Decimal)/ws.cell(5,3).value
ws.cell(11,1,'  标准准入客户通过率')
ws.cell(11,2).value=check_new_data_T.iloc[6,0]
ws.cell(11,3).value=check_new_data_T.iloc[6,1]
ws.cell(12,1,'  降额客户通过率')
ws.cell(12,2).value=check_new_data_T.iloc[5,0]
ws.cell(12,3).value=check_new_data_T.iloc[5,1]
ws.cell(13,1,'老客户指标')
ws.cell(14,1,'老客户放款数')
ws.cell(14,2,fangkuan_old_sum_T.iloc[0,0])
ws.cell(14,3,fangkuan_old_sum_T.iloc[0,1])
ws.cell(15,1,'老客户合同金额')
ws.cell(15,2,fangkuan_old_sum_T.iloc[1,0].astype(int))
ws.cell(15,3,fangkuan_old_sum_T.iloc[1,1].astype(int))
ws.cell(16,1,'老客户合同金额占比')
ws.cell(16,2).value=ws.cell(15,2).value/ws.cell(3,2).value
ws.cell(16,3).value=ws.cell(15,3).value/ws.cell(3,3).value
ws.cell(17,1,'贷后指标(不包括降额客户)')
ws.cell(18,1,'新客户')
ws.cell(19,1,'首逾率')
ws.cell(19,2,dpd1_data_new_user_T.iloc[2,0])
ws.cell(19,3,dpd1_data_new_user_T.iloc[2,1])
ws.cell(20,1,'全款催回率（3天）')
ws.cell(20,2,cuihui_data_new_user_T.iloc[2,0])
ws.cell(20,3,cuihui_data_new_user_T.iloc[2,1])
ws.cell(21,1,'展期催回率（3天）')
ws.cell(21,2,cuihui_data_new_user_T.iloc[3,0])
ws.cell(21,3,cuihui_data_new_user_T.iloc[3,1])
ws.cell(22,1,'dpd7+')
ws.cell(22,2,dpd7_data_new_user_T.iloc[2,0])
ws.cell(22,3,dpd7_data_new_user_T.iloc[2,1])
ws.cell(23,1,'老客户')
ws.cell(24,1,'首逾率')
ws.cell(24,2,dpd1_data_old_user_T.iloc[2,0])
ws.cell(24,3,dpd1_data_old_user_T.iloc[2,1])
ws.cell(25,1,'全款催回率（3天）')
ws.cell(25,2,cuihui_data_old_user_T.iloc[2,0])
ws.cell(25,3,cuihui_data_old_user_T.iloc[2,1])
ws.cell(26,1,'展期催回率（3天）')
ws.cell(26,2,cuihui_data_old_user_T.iloc[3,0])
ws.cell(26,3,cuihui_data_old_user_T.iloc[3,1])
ws.cell(27,1,'dpd7+')
ws.cell(27,2,dpd7_data_old_user_T.iloc[2,0])
ws.cell(27,3,dpd7_data_old_user_T.iloc[2,1])
ws.cell(28,1,'复贷率')
ws.cell(29,1,'分母：放款新客户')
ws.cell(29,2,fudai_sql_clean_T.iloc[4,0])
ws.cell(29,3,fudai_sql_clean_T.iloc[4,1])
ws.cell(30,1,'分母：放款且还款新客户')
ws.cell(30,2,fudai_sql_clean_T.iloc[5,0])
ws.cell(30,3,fudai_sql_clean_T.iloc[5,1])

wb.save(r'C:\Users\Mint\Desktop\weekreport.xlsx')
print('输出完毕')
